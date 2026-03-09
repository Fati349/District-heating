import pulp
import numpy as np
import openpyxl
import math
import os

def solve_instance(filename):
    if not os.path.exists(filename):
        return None
        
    wb = openpyxl.load_workbook(filename, data_only=True)

    # 1. Safely determine N (number of nodes)
    ws_nodes = wb['NodesCord']
    N = 0
    for row in range(1, ws_nodes.max_row + 1):
        val = ws_nodes.cell(row=row, column=1).value
        if val is not None and str(val).strip() != "":
            N += 1

    # 2. Helper functions
    def read_matrix(sheet_name, n=N):
        ws = wb[sheet_name]
        return [[float(ws.cell(row=i+1, column=j+1).value or 0) for j in range(n)] for i in range(n)]

    def read_vector(sheet_name, n=N):
        ws = wb[sheet_name]
        return [float(ws.cell(row=i+1, column=1).value or 0) for i in range(n)]

    def read_scalar(sheet_name):
        return float(wb[sheet_name].cell(row=1, column=1).value)

    source = int(read_scalar('SourceNum')) - 1 

    # 3. Coordinates & Distances
    coords = [(float(wb['NodesCord'].cell(row=i+1, column=1).value), float(wb['NodesCord'].cell(row=i+1, column=2).value)) for i in range(N)]
    l = [[math.sqrt((coords[i][0]-coords[j][0])**2 + (coords[i][1]-coords[j][1])**2) for j in range(N)] for i in range(N)]

    # 4. Load exact parameters (No symmetry modification)
    theta_fix = read_matrix('vfix(thetaijfix)')
    theta_var = read_matrix('vvar(thetaijvar)')
    c_fix = read_scalar('FixedUnitCost')
    c_var = read_matrix('cvar(cijvar)')
    c_heat = read_vector('cheat(ciheat)')
    c_om = read_matrix('com(cijom)')
    c_rev = read_matrix('crev(cijrev)')
    T_flh = read_scalar('Tflh(Tiflh)')
    beta = read_scalar('Betta')
    lam = read_scalar('Lambda')
    alpha = float(wb['Alpha'].cell(row=1, column=1).value)
    d = read_matrix('EdgesDemandPeak(dij)')
    D = read_matrix('EdgesDemandAnnual(Dij)')
    C_max = read_matrix('Cmax(cijmax)')
    Q_max = read_scalar('SourceMaxCap(Qimax)')
    p_umd = read_matrix('pumd(pijumd)')

    nodes = list(range(N))
    edges = [(i, j) for i in nodes for j in nodes if i != j]

    # Initialize Problem
    prob = pulp.LpProblem(f"DistrictHeating_{N}_nodes", pulp.LpMinimize)

    # Decision Variables
    X = {(i,j): pulp.LpVariable(f"X_{i}_{j}", cat='Binary') for (i,j) in edges}
    P_in = {(i,j): pulp.LpVariable(f"Pin_{i}_{j}", lowBound=0) for (i,j) in edges}
    P_out = {(i,j): pulp.LpVariable(f"Pout_{i}_{j}", lowBound=0) for (i,j) in edges}

    # Objective Function Components
    revenue = pulp.lpSum(X[i,j] * D[i][j] * lam * c_rev[i][j] for (i,j) in edges)
    heat_gen = pulp.lpSum(P_in[source,j] for j in nodes if j != source) * T_flh * c_heat[source] / beta
    maintenance = pulp.lpSum(X[i,j] * l[i][j] * c_om[i][j] for (i,j) in edges)
    fixed_inv = alpha * pulp.lpSum(X[i,j] * l[i][j] * c_fix for (i,j) in edges)
    var_inv = alpha * pulp.lpSum(P_in[i,j] * l[i][j] * c_var[i][j] for (i,j) in edges)
    
    # Original, correct unmet penalty logic (avoids double-counting the penalty)
    unmet = pulp.lpSum((1 - X[i,j] - X[j,i]) * D[i][j] * p_umd[i][j] for i in nodes for j in nodes if i < j)

    # Compile Objective
    prob += heat_gen + maintenance + fixed_inv + var_inv + unmet - revenue

    # --- Constraints ---
    prob += pulp.lpSum(X[i,j] for (i,j) in edges) == N - 1

    for i in nodes:
        for j in nodes:
            if i < j:
                prob += X[i,j] + X[j,i] <= 1

    for (i,j) in edges:
        delta = (d[i][j] * beta * lam) + (theta_fix[i][j] * l[i][j])
        prob += P_out[i,j] == P_in[i,j] * (1 - theta_var[i][j] * l[i][j]) - delta * X[i,j]

    for j in nodes:
        if j != source:
            prob += pulp.lpSum(P_out[i,j] for i in nodes if i != j) == \
                   pulp.lpSum(P_in[j,i] for i in nodes if i != j)

    for (i,j) in edges:
        prob += P_in[i,j] <= C_max[i][j] * X[i,j]

    prob += pulp.lpSum(P_in[j, source] for j in nodes if j != source) == 0

    prob += pulp.lpSum(P_in[source,j] for j in nodes if j != source) <= Q_max

    for j in nodes:
        if j != source:
            prob += pulp.lpSum(X[i,j] for i in nodes if i != j) >= 1

    # Simple solver execution
    prob.solve(pulp.PULP_CBC_CMD(msg=0))

    return pulp.value(prob.objective)

if __name__ == '__main__':
    val_small = solve_instance('InputDataEnergySmallInstance.xlsx')
    val_large = solve_instance('InputDataEnergyLargeInstance.xlsx')
    
    print("Optimal solutions")
    
    if val_small is not None:
        print("Small instance")
        print(f"The optimal value for the small instance is {val_small:.8f}.")
        
    if val_large is not None:
        print("Large instance")
        print(f"The optimal value for the large instance is {val_large:.12e}.")